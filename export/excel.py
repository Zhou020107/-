import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import io


def export_to_excel(jobs_df: pd.DataFrame, filters: dict = None) -> bytes:
    """
    导出为格式化的 Excel 文件 (.xlsx)
    返回 bytes 供 Streamlit 下载
    """
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 全部职位
    _write_all_jobs(wb, jobs_df)

    # Sheet 2-5: 分类透视
    if not jobs_df.empty:
        _write_category_sheets(wb, jobs_df)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def _write_all_jobs(wb: Workbook, df: pd.DataFrame):
    """写入全部职位主表"""
    ws = wb.create_sheet("全部职位", 0)

    columns = {
        'title': '职位名称', 'company': '公司', 'location': '工作地点',
        'location_tier': '城市等级', 'salary_min': '薪资下限', 'salary_max': '薪资上限',
        'salary_text': '薪资原文', 'salary_range': '薪资档位',
        'education': '学历要求', 'experience': '经验要求',
        'tech_stack': '技术栈', 'source': '来源网站', 'post_date': '发布日期',
        'job_type': '职位类型', 'description': '职位描述', 'url': '详情链接',
        'is_accept_graduate': '招应届生', 'has_weekend_off': '双休', 'has_insurance': '五险一金',
    }

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")
    cell_font = Font(size=10, name="微软雅黑")
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    link_font = Font(size=10, color="0563C1", underline="single", name="微软雅黑")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Headers
    col_idx = 1
    displayed_cols = []
    for col, header in columns.items():
        if col in df.columns:
            displayed_cols.append(col)
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            col_idx += 1

    # Data
    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = row_idx + 2
        is_alt = row_idx % 2 == 1

        for col_offset, col_name in enumerate(displayed_cols):
            val = row.get(col_name, '')
            if val is None:
                val = ''

            cell = ws.cell(row=excel_row, column=col_offset + 1, value=val)

            if is_alt:
                cell.fill = alt_fill

            cell.font = cell_font
            cell.border = thin_border

            # URL 列特殊处理
            if col_name == 'url' and val:
                cell.hyperlink = str(val)
                cell.font = link_font

            # 数字列格式化
            if col_name in ('salary_min', 'salary_max') and val:
                cell.number_format = '#,##0'

    # Column width auto-fit
    for col_idx, col_name in enumerate(displayed_cols):
        max_len = len(columns.get(col_name, col_name)) * 2
        for row_idx in range(min(50, len(df))):
            val = str(df.iloc[row_idx].get(col_name, ''))
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx + 1)].width = min(max_len + 4, 40)

    # Freeze top row
    ws.freeze_panes = 'A2'
    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(displayed_cols))}{len(df) + 1}"


def _write_category_sheets(wb: Workbook, df: pd.DataFrame):
    """写入分类工作表"""
    # 按学历
    if 'education' in df.columns:
        ws = wb.create_sheet("按学历")
        _write_grouped_sheet(ws, df, 'education', '学历要求')

    # 按薪资
    if 'salary_range' in df.columns:
        ws = wb.create_sheet("按薪资")
        _write_grouped_sheet(ws, df, 'salary_range', '薪资档位')

    # 按城市
    if 'location_tier' in df.columns:
        ws = wb.create_sheet("按城市")
        _write_grouped_sheet(ws, df, 'location_tier', '城市等级')

    # 统计摘要
    ws = wb.create_sheet("统计摘要")
    summary_data = [
        ['统计项目', '数值'],
        ['职位总数', len(df)],
        ['来源网站数', df['source'].nunique() if 'source' in df.columns else 0],
        ['公司数', df['company'].nunique() if 'company' in df.columns else 0],
        ['城市数', df['location'].nunique() if 'location' in df.columns else 0],
        ['导出时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ]
    for row_idx, row_data in enumerate(summary_data):
        for col_idx, val in enumerate(row_data):
            ws.cell(row=row_idx + 1, column=col_idx + 1, value=val)

    # 来源分布
    if 'source' in df.columns:
        ws.append([])
        ws.append(['来源分布:'])
        for source, cnt in df['source'].value_counts().items():
            ws.append([source, cnt])


def _write_grouped_sheet(ws, df: pd.DataFrame, group_col: str, group_name: str):
    """按某列分组写入"""
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11, name="微软雅黑")

    display_cols = ['title', 'company', 'location', 'salary_text', 'education', 'tech_stack', 'source']
    headers = ['职位名称', '公司', '地点', '薪资', '学历', '技术栈', '来源']
    row = 1

    for group_val, group_df in df.groupby(group_col, dropna=False):
        val = group_val if group_val else '未知'
        # Group header
        cell = ws.cell(row=row, column=1, value=f"{group_name}: {val} ({len(group_df)}条)")
        cell.font = Font(bold=True, size=12, name="微软雅黑", color="1F4E79")
        row += 1

        # Column headers
        for col_idx, h in enumerate(headers):
            cell = ws.cell(row=row, column=col_idx + 1, value=h)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        # Data
        for _, job in group_df.iterrows():
            for col_idx, col_name in enumerate(display_cols):
                if col_name in df.columns:
                    val = job.get(col_name, '')
                    ws.cell(row=row, column=col_idx + 1, value=val if val else '')
            row += 1

        row += 1  # blank row between groups
